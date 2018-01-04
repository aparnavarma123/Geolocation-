from Tkinter import *
import tkFileDialog
import tweepy
from tweepy import Stream
from tweepy import OAuthHandler
from tweepy.streaming import StreamListener
import json
__version__='2.4.8'
import openpyxl
from openpyxl import __version__
from openpyxl import Workbook


CONSUMER_KEY = 'dfdBtqrvwZXfzMioBfNCdcqaT'
CONSUMER_SECRET = 'FCf9qxTKcDSNdai34Sg5kltfAxSuts6dMlpUivD4qEI6haEin6'
ACCESS_KEY = '871616161644912640-fQ9MvSnFJWXiUJqh6sliSCZopDVmFhP'
ACCESS_SECRET = 'HTRcAZdMnVnGUQiGhKKGxpAii3adUzRyccqxcbU7nGl3V'


a=''
root = Tk(className ="Twitter Location") #add a root window named Twitter Location
foo1= Label(root,text="Input the Source File ",height=2,width=35) # add a label to root window
def act(): # defines an event function - for click of button
    file = tkFileDialog.askopenfile(parent=root,mode='rb+',title='Choose a file')  
    global a
    a=file.name
    print a
    book = openpyxl.load_workbook(file)
    sheetR = book.active
    i=2
    sheetR.cell(row=1,column=45).value=" User Id "
    while(sheetR.cell(row=i,column=3) is not None):
        a1 = sheetR.cell(row=i,column=3)
        s=a1.value
        if(s is  None):
            break;
        s=s.partition('m/')[-1].rpartition('/s')[0]
        sheetR.cell(row=i,column=45).value=s
        i=i+1
    book.save(a)
    root.destroy()
    print
    print 'IN PROCESS'
foo = Button(root,text="Browse File",width=15,height=1,command=act)
foo1.pack()
foo.pack()
root.mainloop()

book = openpyxl.load_workbook(a)
sheet1=book.active
sheet1.cell(row=1,column=46).value='User Name'
sheet1.cell(row=1,column=47).value='Full Name'
sheet1.cell(row=1,column=48).value='Bio'
sheet1.cell(row=1,column=49).value='Location'
sheet1.cell(row=1,column=50).value='Tweets\' Location'
sheet1.cell(row=1,column=51).value='Language Setting'
book.save(a)


class TweetListener(StreamListener):
    def on_data(self, data):
        print data
        return True
    def on_error(self, status):
        print status

auth = OAuthHandler(CONSUMER_KEY,CONSUMER_SECRET)
api = tweepy.API(auth)
auth.set_access_token(ACCESS_KEY, ACCESS_SECRET)
twitterStream = Stream(auth,TweetListener())


i=2
j=1
def userInfo(IdString):
    global j
    user = api.get_user(IdString)
    sheet1.cell(row=j+1,column=46).value=user.screen_name
    sheet1.cell(row=j+1,column=47).value=user.name
    if (user.description):
        sheet1.cell(row=j+1,column=48).value=user.description
    else:
        sheet1.cell(row=j+1,column=48).value=" "
    if (user.location):
        sheet1.cell(row=j+1,column=49).value=user.location
    else:
        sheet1.cell(row=j+1,column=49).value=' '
    if(user.lang):
        sheet1.cell(row=j+1,column=51).value=user.lang
    else:
        sheet1.cell(row=j+1,column=51).value=' '
    book.save(a)

def get_all_tweets(screen_name):
    auth = OAuthHandler(CONSUMER_KEY,CONSUMER_SECRET)
    api = tweepy.API(auth)
    auth.set_access_token(ACCESS_KEY, ACCESS_SECRET)
    alltweets = []   #initialize a list to hold all the tweepy Tweets 
    new_tweets = api.user_timeline(screen_name = screen_name,count=200)
    alltweets.extend(new_tweets)
    oldest = alltweets[-1].id - 1
    
    while len(new_tweets) > 0 and len(alltweets)<600:
        new_tweets = api.user_timeline(screen_name = screen_name,count=200,max_id=oldest)
        alltweets.extend(new_tweets)
        oldest = alltweets[-1].id - 1
       
    file = open('tweet.json', 'wb')    
    for status in alltweets:
        json.dump(status._json,file,sort_keys = True,indent =1)
    file.close()
k=1
def wordExtMat():
    
    f=0 #flag for location
    file = open('JSONtoWord.txt', 'w')
    fo=open ('tweet.json','r')
    char= fo.read()
    word=''
    for i in range(len(char)):
        if(char[i].isalpha()):
            word=word+char[i]
            continue
        if (word is not ''):       
            file.write(word+'\n') 
        word=''
    fo.close()
    file.close()
    fo3=open('JSONtoWord.txt', 'r')
    fo2=open('city.txt','r')
    char3=fo3.readlines()
    char2=fo2.readlines()
    global k
    for i in char3:
        for j in char2:
            if(j==i):
                f=1
                sheet1.cell(row=k+1,column=50).value=j
                break
        if(f==1):
            break
    if (f==0):
        sheet1.cell(row=k+1,column=50).value=" "
    book.save(a)



num_rows = sheet1.max_row
for y in range(num_rows-1):
    a1=sheet1.cell(row=y+2,column=45).value
    print a1
    userInfo(a1)
    user = api.get_user(a1)
    get_all_tweets(user.screen_name)
    wordExtMat()    
    k=k+1
    i=i+1
    j=j+1
    print "Done"
book.save(a)


root = Tk(className ="Twitter Location") #add a root window named
foo1= Label(root,text="Output File is Stored ",height=2,width=35) # add a label to root window
def openB():
    from os import startfile
    global a
    startfile(a)
    root.destroy()
    

foo2=Button(root,text="Show Output File",width=15,height=1,command=openB)
foo1.pack()
foo2.pack()
root.mainloop()

